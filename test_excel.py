#!/usr/bin/env python3
"""
Test Excel file reading to identify segfault source
"""

import sys
import traceback
import os

def test_pandas_excel():
    """Test pandas Excel reading"""
    try:
        print("Testing pandas Excel reading...")
        import pandas as pd
        
        # Test file
        test_file = "incremental_tasks/task1_train.xlsx"
        
        if not os.path.exists(test_file):
            print(f"Test file {test_file} not found!")
            return False
            
        print(f"Reading {test_file}...")
        
        # Try different engines
        engines = ['openpyxl', 'xlrd', None]
        
        for engine in engines:
            try:
                print(f"Trying engine: {engine}")
                if engine:
                    data = pd.read_excel(test_file, engine=engine)
                else:
                    data = pd.read_excel(test_file)
                    
                print(f"✓ Success with engine {engine}")
                print(f"Shape: {data.shape}")
                print(f"Columns: {list(data.columns)}")
                
                # Test data access
                if 'description' in data.columns:
                    print(f"First description: {data['description'].iloc[0][:100]}...")
                    
                return True
                
            except Exception as e:
                print(f"✗ Failed with engine {engine}: {e}")
                continue
                
        print("All engines failed!")
        return False
        
    except Exception as e:
        print(f"Pandas test failed: {e}")
        traceback.print_exc()
        return False

def test_openpyxl_direct():
    """Test openpyxl directly"""
    try:
        print("\nTesting openpyxl directly...")
        import openpyxl
        
        test_file = "incremental_tasks/task1_train.xlsx"
        
        print(f"Opening {test_file} with openpyxl...")
        wb = openpyxl.load_workbook(test_file)
        ws = wb.active
        
        print(f"✓ Workbook loaded")
        print(f"Sheet name: {ws.title}")
        print(f"Max row: {ws.max_row}")
        print(f"Max col: {ws.max_column}")
        
        # Read first few cells
        for row in range(1, min(6, ws.max_row + 1)):
            for col in range(1, min(4, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                print(f"Cell ({row},{col}): {str(cell_value)[:50]}...")
                
        wb.close()
        return True
        
    except Exception as e:
        print(f"OpenPyXL test failed: {e}")
        traceback.print_exc()
        return False

def main():
    """Main test function"""
    print("=== Excel Reading Test ===")
    
    # Test pandas
    pandas_ok = test_pandas_excel()
    
    # Test openpyxl directly
    openpyxl_ok = test_openpyxl_direct()
    
    if pandas_ok or openpyxl_ok:
        print("\n✅ At least one method works!")
        return True
    else:
        print("\n❌ All methods failed!")
        return False

if __name__ == "__main__":
    try:
        success = main()
        sys.exit(0 if success else 1)
    except KeyboardInterrupt:
        print("\n⚠ Test interrupted")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Fatal error: {e}")
        traceback.print_exc()
        sys.exit(1)